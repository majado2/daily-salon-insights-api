from collections import defaultdict
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any, Literal, cast
from urllib.parse import quote
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select

from app.api.admin_master import paginate
from app.api.deps import AdminUser, DbSession, envelope
from app.core.errors import BusinessError
from app.core.time import iso_utc
from app.models.entities import Branch
from app.services.business import add_audit
from app.services.reports import (
    ReconciliationFilter,
    aggregate_rows,
    employee_sale_rows,
    ranking_rows,
    report_summary,
    settlement_rows,
)

router = APIRouter(prefix="/admin/reports", tags=["Reports"])
RIYADH = ZoneInfo("Asia/Riyadh")
VALID_RECONCILIATION = {"matched", "surplus", "shortage", "provisional"}


def validate_dates(date_from: date, date_to: date) -> None:
    if date_to < date_from:
        raise BusinessError(
            422, "INVALID_DATE_RANGE", "The end date cannot precede the start date."
        )


def reconciliation_value(value: str | None) -> ReconciliationFilter | None:
    if value is None:
        return None
    if value not in VALID_RECONCILIATION:
        raise BusinessError(
            422, "INVALID_RECONCILIATION_FILTER", "The reconciliation filter is invalid."
        )
    return cast(ReconciliationFilter, value)


def serialize_times(row: dict[str, object]) -> dict[str, object]:
    copy = dict(row)
    copy["cashier_updated_at"] = iso_utc(row.get("cashier_updated_at"))  # type: ignore[arg-type]
    copy["admin_updated_at"] = iso_utc(row.get("admin_updated_at"))  # type: ignore[arg-type]
    return copy


@router.get("/settlements")
def settlements_report(
    request: Request,
    user: AdminUser,
    db: DbSession,
    date_from: date,
    date_to: date,
    branch_id: int | None = None,
    employee_id: int | None = None,
    reconciliation_filter: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
) -> dict[str, object]:
    del user
    validate_dates(date_from, date_to)
    rows = settlement_rows(
        db,
        date_from=date_from,
        date_to=date_to,
        branch_id=branch_id,
        employee_id=employee_id,
        reconciliation_filter=reconciliation_value(reconciliation_filter),
    )
    page_rows, meta = paginate(rows, page, page_size)
    meta["summary"] = report_summary(rows)  # type: ignore[assignment]
    if employee_id:
        meta["scope_note"] = "branch_totals_not_employee_attributed"  # type: ignore[assignment]
    return envelope(request, [serialize_times(row) for row in page_rows], meta=meta)


@router.get("/employee-sales")
def employee_sales_report(
    request: Request,
    user: AdminUser,
    db: DbSession,
    date_from: date,
    date_to: date,
    branch_id: int | None = None,
    employee_id: int | None = None,
    reconciliation_filter: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=100),
) -> dict[str, object]:
    del user
    validate_dates(date_from, date_to)
    rows = employee_sale_rows(
        db,
        date_from=date_from,
        date_to=date_to,
        branch_id=branch_id,
        employee_id=employee_id,
        reconciliation_filter=reconciliation_value(reconciliation_filter),
    )
    page_rows, meta = paginate(rows, page, page_size)
    sales_total = sum((Decimal(str(row["sales_total"])) for row in rows), Decimal("0.00"))
    meta["summary"] = {
        "service_count": sum(int(row["service_count"]) for row in rows),
        "sales_total": f"{sales_total:.2f}",
    }  # type: ignore[assignment]
    return envelope(request, page_rows, meta=meta)


@router.get("/aggregates")
def aggregates_report(
    request: Request,
    user: AdminUser,
    db: DbSession,
    date_from: date,
    date_to: date,
    group_by: Literal["day", "month", "year", "branch"],
    branch_id: int | None = None,
    employee_id: int | None = None,
    reconciliation_filter: str | None = None,
) -> dict[str, object]:
    del user
    validate_dates(date_from, date_to)
    rows = settlement_rows(
        db,
        date_from=date_from,
        date_to=date_to,
        branch_id=branch_id,
        employee_id=employee_id,
        reconciliation_filter=reconciliation_value(reconciliation_filter),
    )
    return envelope(request, aggregate_rows(rows, group_by), meta={"summary": report_summary(rows)})


@router.get("/rankings")
def rankings_report(
    request: Request,
    user: AdminUser,
    db: DbSession,
    entity: Literal["employee", "branch"],
    date_from: date,
    date_to: date,
    branch_id: int | None = None,
    limit: int = Query(10, ge=1, le=100),
) -> dict[str, object]:
    del user
    validate_dates(date_from, date_to)
    return envelope(
        request,
        ranking_rows(
            db,
            entity=entity,
            date_from=date_from,
            date_to=date_to,
            branch_id=branch_id,
            limit=limit,
        ),
    )


def excel_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    aware = value.replace(tzinfo=UTC) if value.tzinfo is None else value
    return aware.astimezone(RIYADH).replace(tzinfo=None)


def report_dates(date_from: date, date_to: date) -> list[date]:
    return [date_from + timedelta(days=offset) for offset in range((date_to - date_from).days + 1)]


def employee_groups(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[int, dict[str, Any]] = {}
    for row in rows:
        employee = row["employee"]
        employee_id = int(employee["id"])
        group = groups.setdefault(
            employee_id,
            {
                "employee": employee,
                "branches": set(),
                "daily": {},
            },
        )
        group["branches"].add(row["branch"]["name"])
        group["daily"][row["work_date"]] = row
    return sorted(
        groups.values(),
        key=lambda item: (
            str(item["employee"]["name"]).casefold(),
            str(item["employee"]["employee_code"]),
        ),
    )


def append_employee_matrix(
    workbook: Workbook,
    rows: list[dict[str, Any]],
    dates: list[date],
    *,
    metric: Literal["sales", "services"],
    arabic: bool,
) -> Worksheet:
    title = (
        "مصفوفة مبيعات العاملات"
        if arabic and metric == "sales"
        else "مصفوفة خدمات العاملات"
        if arabic
        else "Employee sales matrix"
        if metric == "sales"
        else "Employee services matrix"
    )
    sheet = cast(Worksheet, workbook.create_sheet(title))
    headers: list[str | int] = (
        ["رقم العاملة", "العاملة", "الفروع"]
        if arabic
        else ["Employee code", "Employee", "Branches"]
    )
    headers.extend(day.day for day in dates)
    headers.append("الإجمالي" if arabic else "Total")
    sheet.append(headers)
    for group in employee_groups(rows):
        values: list[int | float | None] = []
        for day in dates:
            row = group["daily"].get(day.isoformat())
            if row is None:
                values.append(None)
            elif metric == "sales":
                values.append(float(row["sales_total"]))
            else:
                values.append(int(row["service_count"]))
        present = [value for value in values if value is not None]
        total: int | float = sum(present) if present else 0
        employee = group["employee"]
        sheet.append(
            [
                employee["employee_code"],
                employee["name"],
                "، ".join(sorted(group["branches"])),
                *values,
                total,
            ]
        )
    if metric == "sales":
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=sheet.max_column):
            for cell in row:
                cell.number_format = "#,##0.00"
    return sheet


def append_employee_monthly_summary(
    workbook: Workbook, rows: list[dict[str, Any]], *, arabic: bool
) -> Worksheet:
    sheet = cast(
        Worksheet,
        workbook.create_sheet("ملخص العاملات الشهري" if arabic else "Employee monthly summary"),
    )
    sheet.append(
        [
            "رقم العاملة",
            "العاملة",
            "الفروع",
            "أيام العمل",
            "إجمالي الخدمات",
            "إجمالي المبيعات",
            "متوسط المبيعات اليومي",
        ]
        if arabic
        else [
            "Employee code",
            "Employee",
            "Branches",
            "Work days",
            "Total services",
            "Total sales",
            "Average daily sales",
        ]
    )
    for group in employee_groups(rows):
        daily_rows = list(group["daily"].values())
        work_days = len(daily_rows)
        services = sum(int(row["service_count"]) for row in daily_rows)
        sales = sum((Decimal(str(row["sales_total"])) for row in daily_rows), Decimal("0.00"))
        employee = group["employee"]
        sheet.append(
            [
                employee["employee_code"],
                employee["name"],
                "، ".join(sorted(group["branches"])),
                work_days,
                services,
                float(sales),
                float(sales / work_days) if work_days else 0,
            ]
        )
    for row in sheet.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = "#,##0.00"
    return sheet


def append_branch_daily_matrix(
    workbook: Workbook,
    rows: list[dict[str, Any]],
    dates: list[date],
    *,
    arabic: bool,
    branch_names: list[str],
    single_branch: bool,
    date_from: date,
    date_to: date,
) -> Worksheet:
    sheet = cast(
        Worksheet,
        workbook.create_sheet("ملخص الفروع اليومي" if arabic else "Branch daily matrix"),
    )
    scope_value = (
        ("فرع واحد" if single_branch else "جميع الفروع")
        if arabic
        else ("Single branch" if single_branch else "All branches")
    )
    sheet.append(
        [
            "نطاق التقرير" if arabic else "Report scope",
            scope_value,
        ]
    )
    sheet.append(
        [
            "الفروع المشمولة" if arabic else "Included branches",
            "، ".join(branch_names)
            if branch_names
            else ("لا توجد فروع" if arabic else "No branches"),
        ]
    )
    sheet.append(
        [
            "الفترة" if arabic else "Period",
            f"{date_from.isoformat()} - {date_to.isoformat()}",
        ]
    )
    sheet.append([])
    sheet.append(
        [
            "البند" if arabic else "Metric",
            *(day.strftime("%d/%m") for day in dates),
            "الإجمالي" if arabic else "Total",
        ]
    )
    by_date: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_date[row["work_date"]].append(row)
    metrics = [
        ("إجمالي الكاش" if arabic else "Cash total", "cash_total"),
        ("إجمالي البنك" if arabic else "Bank total", "bank_total"),
        ("إجمالي المبيعات" if arabic else "Employee sales", "employees_sales_total"),
        ("الفروقات" if arabic else "Differences", "difference_amount"),
    ]
    for label, field in metrics:
        values: list[float | None] = []
        for day in dates:
            source = by_date.get(day.isoformat(), [])
            if field == "difference_amount":
                raw_values = [
                    row["reconciliation"]["difference_amount"]
                    for row in source
                    if row["reconciliation"]["difference_amount"] is not None
                ]
            else:
                raw_values = [row[field] for row in source if row[field] is not None]
            values.append(sum(float(value) for value in raw_values) if raw_values else None)
        present = [value for value in values if value is not None]
        sheet.append([label, *values, sum(present) if present else 0])
    for row_number in range(1, 4):
        sheet.cell(row=row_number, column=1).font = Font(bold=True)
    for worksheet_row in sheet.iter_rows(min_row=6, min_col=2, max_col=sheet.max_column):
        for cell in worksheet_row:
            cell.number_format = "#,##0.00"
    return sheet


def configure_excel_sheet(
    sheet: Worksheet,
    *,
    arabic: bool,
    freeze_panes: str,
    minimum_width: int = 12,
    header_row: int = 1,
) -> None:
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = freeze_panes
    sheet.auto_filter.ref = (
        f"A{header_row}:{sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate}"
    )
    sheet.sheet_view.rightToLeft = arabic
    for column in sheet.columns:
        width = min(35, max(minimum_width, max(len(str(cell.value or "")) for cell in column) + 2))
        first_cell = cast(Cell, column[0])
        sheet.column_dimensions[first_cell.column_letter].width = width


@router.get("/export.xlsx")
def export_report(
    request: Request,
    user: AdminUser,
    db: DbSession,
    date_from: date,
    date_to: date,
    branch_id: int | None = None,
    employee_id: int | None = None,
    reconciliation_filter: str | None = None,
) -> StreamingResponse:
    validate_dates(date_from, date_to)
    filter_value = reconciliation_value(reconciliation_filter)
    settlements = settlement_rows(
        db,
        date_from=date_from,
        date_to=date_to,
        branch_id=branch_id,
        employee_id=employee_id,
        reconciliation_filter=filter_value,
    )
    employee_sales = employee_sale_rows(
        db,
        date_from=date_from,
        date_to=date_to,
        branch_id=branch_id,
        employee_id=employee_id,
        reconciliation_filter=filter_value,
    )
    if branch_id is None:
        included_branch_names = list(db.scalars(select(Branch.name).order_by(Branch.name)))
    else:
        selected_branch = db.get(Branch, branch_id)
        included_branch_names = [selected_branch.name] if selected_branch else []
    arabic = request.headers.get("Accept-Language", "ar").lower().startswith("ar")
    workbook = Workbook()
    summary_sheet = cast(Worksheet, workbook.active)
    summary_sheet.title = "ملخص التسويات" if arabic else "Settlement summary"
    detail_sheet = cast(
        Worksheet, workbook.create_sheet("تفاصيل العاملات" if arabic else "Employee details")
    )
    summary_headers = (
        [
            "التاريخ",
            "الفرع",
            "عدد الخدمات",
            "مبيعات العاملات",
            "الكاش",
            "البنك",
            "دخل الفرع",
            "حالة المطابقة",
            "النتيجة",
            "الفرق",
            "الملاحظة",
            "منشئ التسوية",
            "آخر كاشيرة معدلة",
            "وقت تعديل الكاشيرة",
            "آخر مدير معدل",
            "وقت تعديل المدير",
        ]
        if arabic
        else [
            "Date",
            "Branch",
            "Services",
            "Employee sales",
            "Cash",
            "Bank",
            "Branch income",
            "Reconciliation",
            "Result",
            "Difference",
            "Note",
            "Created by",
            "Last cashier editor",
            "Cashier updated at",
            "Last admin editor",
            "Admin updated at",
        ]
    )
    detail_headers = (
        [
            "التاريخ",
            "الفرع",
            "رقم العاملة",
            "العاملة",
            "عدد الخدمات",
            "مبيعات العاملة",
            "منشئ السجل",
            "آخر كاشيرة معدلة",
            "آخر مدير معدل",
        ]
        if arabic
        else [
            "Date",
            "Branch",
            "Employee code",
            "Employee",
            "Services",
            "Employee sales",
            "Created by",
            "Last cashier editor",
            "Last admin editor",
        ]
    )
    summary_sheet.append(summary_headers)
    detail_sheet.append(detail_headers)
    for row in settlements:
        reconciliation = row["reconciliation"]
        summary_sheet.append(
            [
                date.fromisoformat(row["work_date"]),
                row["branch"]["name"],
                row["service_count"],
                float(row["employees_sales_total"] or 0),
                float(row["cash_total"] or 0) if row["cash_total"] is not None else None,
                float(row["bank_total"] or 0) if row["bank_total"] is not None else None,
                float(row["branch_income_total"] or 0)
                if row["branch_income_total"] is not None
                else None,
                reconciliation["status"],
                "مؤقتة"
                if arabic and reconciliation["is_provisional"]
                else "نهائية"
                if arabic
                else "Provisional"
                if reconciliation["is_provisional"]
                else "Final",
                float(reconciliation["difference_amount"] or 0)
                if reconciliation["difference_amount"] is not None
                else None,
                row["reconciliation_note"],
                row["created_by"]["display_name"] if row["created_by"] else None,
                row["cashier_updated_by"]["display_name"] if row["cashier_updated_by"] else None,
                excel_datetime(row["cashier_updated_at"]),
                row["admin_updated_by"]["display_name"] if row["admin_updated_by"] else None,
                excel_datetime(row["admin_updated_at"]),
            ]
        )
    for row in employee_sales:
        detail_sheet.append(
            [
                date.fromisoformat(row["work_date"]),
                row["branch"]["name"],
                row["employee"]["employee_code"],
                row["employee"]["name"],
                row["service_count"],
                float(row["sales_total"]),
                row["created_by"]["display_name"] if row["created_by"] else None,
                row["cashier_updated_by"]["display_name"] if row["cashier_updated_by"] else None,
                row["admin_updated_by"]["display_name"] if row["admin_updated_by"] else None,
            ]
        )
    if settlements:
        summary = report_summary(settlements)
        summary_sheet.append(
            [
                "الإجمالي" if arabic else "Total",
                None,
                summary["service_count"],
                float(summary["employees_sales_total"]),
                float(summary["cash_total"]),
                float(summary["bank_total"]),
                float(summary["branch_income_total"]),
                None,
                None,
                float(summary["difference_amount"]),
                None,
                None,
                None,
                None,
                None,
                None,
            ]
        )
        for cell in summary_sheet[summary_sheet.max_row]:
            cell.font = Font(bold=True)

    dates = report_dates(date_from, date_to)
    employee_sales_matrix = append_employee_matrix(
        workbook, employee_sales, dates, metric="sales", arabic=arabic
    )
    employee_services_matrix = append_employee_matrix(
        workbook, employee_sales, dates, metric="services", arabic=arabic
    )
    employee_monthly_summary = append_employee_monthly_summary(
        workbook, employee_sales, arabic=arabic
    )
    branch_daily_matrix = append_branch_daily_matrix(
        workbook,
        settlements,
        dates,
        arabic=arabic,
        branch_names=included_branch_names,
        single_branch=branch_id is not None,
        date_from=date_from,
        date_to=date_to,
    )
    configure_excel_sheet(summary_sheet, arabic=arabic, freeze_panes="A2")
    configure_excel_sheet(detail_sheet, arabic=arabic, freeze_panes="A2")
    configure_excel_sheet(employee_sales_matrix, arabic=arabic, freeze_panes="D2", minimum_width=8)
    configure_excel_sheet(
        employee_services_matrix, arabic=arabic, freeze_panes="D2", minimum_width=8
    )
    configure_excel_sheet(employee_monthly_summary, arabic=arabic, freeze_panes="A2")
    configure_excel_sheet(
        branch_daily_matrix,
        arabic=arabic,
        freeze_panes="B6",
        minimum_width=8,
        header_row=5,
    )
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    add_audit(
        db,
        actor=user,
        event_type="report.exported",
        entity_type="report",
        entity_id=None,
        branch_id=branch_id,
        new_values={
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "branch_id": branch_id,
            "employee_id": employee_id,
            "reconciliation_filter": filter_value,
            "settlement_rows": len(settlements),
            "employee_rows": len(employee_sales),
        },
    )
    db.commit()
    filename = (
        f"تقرير-المبيعات-{date_from.isoformat()}-{date_to.isoformat()}.xlsx"
        if arabic
        else f"sales-report-{date_from.isoformat()}-{date_to.isoformat()}.xlsx"
    )
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}",
        "Cache-Control": "no-store",
        "X-Request-ID": request.state.request_id,
    }
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
