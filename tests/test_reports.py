from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.models.entities import Employee
from tests.conftest import cashier_headers, login


def complete_day(
    client: TestClient, employees: list[Employee], csrf: str, work_date: object
) -> None:
    for index, employee in enumerate(employees, start=1):
        response = client.put(
            f"/api/v1/cashier/today/employees/{employee.id}/sale",
            json={"service_count": index, "sales_total": f"{index * 100}.00"},
            headers=cashier_headers(csrf, work_date),  # type: ignore[arg-type]
        )
        assert response.status_code == 200
    response = client.put(
        "/api/v1/cashier/today/settlement",
        json={"cash_total": "150.00", "bank_total": "150.00"},
        headers=cashier_headers(csrf, work_date),  # type: ignore[arg-type]
    )
    assert response.status_code == 200


def test_reports_keep_branch_settlement_separate_from_employee_sales(
    client: TestClient, seeded: dict[str, object]
) -> None:
    employees = seeded["employees"]
    assert isinstance(employees, list)
    cashier_csrf = login(client, "0500000101")
    complete_day(client, employees, cashier_csrf, seeded["today"])
    client.cookies.clear()
    login(client, "0500000001")
    today = seeded["today"]
    settlement_response = client.get(
        f"/api/v1/admin/reports/settlements?date_from={today}&date_to={today}"
    )
    assert settlement_response.status_code == 200, settlement_response.text
    settlement = settlement_response.json()
    assert settlement["meta"]["summary"] == {
        "service_count": 3,
        "employees_sales_total": "300.00",
        "cash_total": "150.00",
        "bank_total": "150.00",
        "branch_income_total": "300.00",
        "difference_amount": "0.00",
    }
    employee_response = client.get(
        f"/api/v1/admin/reports/employee-sales?date_from={today}&date_to={today}"
    )
    assert employee_response.status_code == 200
    employee_rows = employee_response.json()["data"]
    assert len(employee_rows) == 2
    assert all("cash_total" not in row and "bank_total" not in row for row in employee_rows)
    assert sum(float(row["sales_total"]) for row in employee_rows) == 300.0


def test_excel_has_summary_and_employee_detail_without_cash_columns(
    client: TestClient, seeded: dict[str, object]
) -> None:
    employees = seeded["employees"]
    assert isinstance(employees, list)
    cashier_csrf = login(client, "0500000101")
    complete_day(client, employees, cashier_csrf, seeded["today"])
    client.cookies.clear()
    login(client, "0500000001")
    today = seeded["today"]
    response = client.get(
        f"/api/v1/admin/reports/export.xlsx?date_from={today}&date_to={today}",
        headers={"Accept-Language": "en"},
    )
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == [
        "Settlement summary",
        "Employee details",
        "Employee sales matrix",
        "Employee services matrix",
        "Employee monthly summary",
        "Branch daily matrix",
    ]
    detail_headers = [cell.value for cell in workbook["Employee details"][1]]
    assert "Cash" not in detail_headers
    assert "Bank" not in detail_headers
    summary = workbook["Settlement summary"]
    headers = [cell.value for cell in summary[1]]
    values = [cell.value for cell in summary[2]]
    row = dict(zip(headers, values, strict=True))
    assert row["Employee sales"] == 300
    assert row["Cash"] == 150
    assert row["Bank"] == 150
    assert row["Branch income"] == 300

    employee_day_label = seeded["today"].day  # type: ignore[union-attr]
    sales_matrix = workbook["Employee sales matrix"]
    sales_headers = [cell.value for cell in sales_matrix[1]]
    sales_by_employee = {
        values["Employee code"]: values
        for values in (
            dict(zip(sales_headers, [cell.value for cell in row], strict=True))
            for row in sales_matrix.iter_rows(min_row=2)
        )
    }
    first_employee = sales_by_employee["E0001"]
    assert first_employee["Employee code"] == "E0001"
    assert first_employee[employee_day_label] == 100
    assert first_employee["Total"] == 100

    services_matrix = workbook["Employee services matrix"]
    service_headers = [cell.value for cell in services_matrix[1]]
    services_by_employee = {
        values["Employee code"]: values
        for values in (
            dict(zip(service_headers, [cell.value for cell in row], strict=True))
            for row in services_matrix.iter_rows(min_row=2)
        )
    }
    second_employee = services_by_employee["E0002"]
    assert second_employee["Employee code"] == "E0002"
    assert second_employee[employee_day_label] == 2

    branch_matrix = workbook["Branch daily matrix"]
    assert branch_matrix["A1"].value == "Report scope"
    assert branch_matrix["B1"].value == "All branches"
    assert branch_matrix["A2"].value == "Included branches"
    assert seeded["branch"].name in branch_matrix["B2"].value  # type: ignore[union-attr,operator]
    assert branch_matrix["A3"].value == "Period"
    branch_day_label = seeded["today"].strftime("%d/%m")  # type: ignore[union-attr]
    branch_headers = [cell.value for cell in branch_matrix[5]]
    branch_rows = {
        values[0]: dict(zip(branch_headers, values, strict=True))
        for values in ([cell.value for cell in row] for row in branch_matrix.iter_rows(min_row=6))
    }
    assert branch_rows["Cash total"][branch_day_label] == 150
    assert branch_rows["Bank total"][branch_day_label] == 150
    assert branch_rows["Employee sales"][branch_day_label] == 300
    assert branch_rows["Differences"][branch_day_label] == 0

    branch = seeded["branch"]
    single_response = client.get(
        f"/api/v1/admin/reports/export.xlsx?date_from={today}&date_to={today}&branch_id={branch.id}",  # type: ignore[union-attr]
        headers={"Accept-Language": "en"},
    )
    assert single_response.status_code == 200
    single_workbook = load_workbook(BytesIO(single_response.content), data_only=True)
    single_branch_matrix = single_workbook["Branch daily matrix"]
    assert single_branch_matrix["B1"].value == "Single branch"
    assert single_branch_matrix["B2"].value == branch.name  # type: ignore[union-attr]

    arabic_response = client.get(
        f"/api/v1/admin/reports/export.xlsx?date_from={today}&date_to={today}&branch_id={branch.id}",  # type: ignore[union-attr]
        headers={"Accept-Language": "ar"},
    )
    assert arabic_response.status_code == 200
    arabic_workbook = load_workbook(BytesIO(arabic_response.content), data_only=True)
    arabic_branch_matrix = arabic_workbook["ملخص الفروع اليومي"]
    assert arabic_branch_matrix["A1"].value == "نطاق التقرير"
    assert arabic_branch_matrix["B1"].value == "فرع واحد"
    assert arabic_branch_matrix["A2"].value == "الفروع المشمولة"
    assert arabic_branch_matrix["B2"].value == branch.name  # type: ignore[union-attr]
    arabic_sales_headers = [cell.value for cell in arabic_workbook["مصفوفة مبيعات العاملات"][1]]
    assert employee_day_label in arabic_sales_headers


def test_employee_details_include_profile_metadata(
    client: TestClient, seeded: dict[str, object]
) -> None:
    employee = seeded["employees"][0]  # type: ignore[index]
    assert isinstance(employee, Employee)
    login(client, "0500000001")
    response = client.get(f"/api/v1/admin/employees/{employee.id}")
    assert response.status_code == 200
    profile = response.json()["data"]
    assert profile["employee_code"] == "E0001"
    assert profile["branch"]["id"] == seeded["branch"].id  # type: ignore[union-attr]
    assert profile["effective_from"] is not None
    assert profile["created_at"].endswith("Z")
    assert profile["updated_at"].endswith("Z")


def test_reconciliation_filter_distinguishes_provisional_results(
    client: TestClient, seeded: dict[str, object]
) -> None:
    employee = seeded["employees"][0]  # type: ignore[index]
    assert isinstance(employee, Employee)
    cashier_csrf = login(client, "0500000101")
    client.put(
        f"/api/v1/cashier/today/employees/{employee.id}/sale",
        json={"service_count": 5, "sales_total": "500.00"},
        headers=cashier_headers(cashier_csrf, seeded["today"]),  # type: ignore[arg-type]
    )
    client.put(
        "/api/v1/cashier/today/settlement",
        json={"cash_total": "250.00", "bank_total": "250.00"},
        headers=cashier_headers(cashier_csrf, seeded["today"]),  # type: ignore[arg-type]
    )
    client.cookies.clear()
    login(client, "0500000001")
    today = seeded["today"]
    provisional = client.get(
        f"/api/v1/admin/reports/settlements?date_from={today}&date_to={today}&reconciliation_filter=provisional"
    )
    matched = client.get(
        f"/api/v1/admin/reports/settlements?date_from={today}&date_to={today}&reconciliation_filter=matched"
    )
    assert len(provisional.json()["data"]) == 1
    assert matched.json()["data"] == []

    aggregate = client.get(
        f"/api/v1/admin/reports/aggregates?date_from={today}&date_to={today}&group_by=day"
    )
    assert aggregate.status_code == 200
    counts = aggregate.json()["data"][0]
    assert counts["provisional_days"] == 1
    assert counts["matched_days"] == 0
